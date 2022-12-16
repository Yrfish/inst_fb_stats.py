from openpyxl.styles import Side, Border
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from auth_data import login, password, email, pas
import time
import random
import locale
from datetime import date
from openpyxl import Workbook
import re


options = webdriver.ChromeOptions()

# Убираем уведомление на фейсбуке

prefs = {"profile.default_content_setting_values.notifications": 2}
options.add_experimental_option("prefs", prefs)
url = "https://www.instagram.com/"
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=options)

try:
    driver.get(url=url)  # open inst
    time.sleep(random.randrange(3, 5))
    # вводим логин инстаграма
    username_input = driver.find_element(By.NAME, "username")
    username_input.clear()
    username_input.send_keys(login)

    time.sleep(random.randrange(3, 5))

    password_input = driver.find_element(By.NAME, "password")
    password_input.clear()
    password_input.send_keys(password)

    time.sleep(random.randrange(3, 5))

    password_input.send_keys(Keys.ENTER)

    time.sleep(random.randrange(5, 15))

    driver.get("https://www.instagram.com/yaroslav_fishing_/")

    time.sleep(random.randrange(2, 5))

    hrefs = driver.find_elements(By.TAG_NAME, "a")
    post_urls = [i.get_attribute("href") for i in hrefs if "/p/" in i.get_attribute("href")]
    inst_data = list()

    for url in post_urls[2:10]:
        driver.get(url=url)
        time.sleep(random.randrange(4, 8))
        try:
            btns = driver.find_element(By.XPATH,
                                       "//BUTTON[@class='_acan _acao _acas'][text()='Переглянути статистику']").click()  # нажимаем на кнопку статистика в публикации
        except:
            print("Для рілс додаш в ручну :)")
            post_info = {"Link": url, "coment": "-", "likes": "-", "coment_count": "-", "coverage": "-"}
            inst_data.append(post_info)
            print(url)
            continue

        time.sleep(3)
        likes = driver.find_element(By.XPATH,
                                    "//SPAN[@data-bloks-name='bk.components.Text']").text
        coment_count = driver.find_element(By.XPATH,
                                           "//SPAN[@data-bloks-name='bk.components.Text']").text
        coverage = driver.find_element(By.XPATH,
                                       "//SPAN[@data-bloks-name='bk.components.Text']").text
        post_info = {"Link": url, "coment": "-", "likes": likes, "coment_count": coment_count, "coverage": coverage}
        print(f"url = {url}, like = {likes}, comm = {coment_count}, coverage = {coverage}")
        inst_data.append(post_info)
        time.sleep(3)


    # начинаем парсить фейсбук
    driver.get(url="https://www.facebook.com")  # going to facebook general page
    time.sleep(5)
    # Находим поля ввода пароля и логина
    input_email = driver.find_element(By.CSS_SELECTOR, "input[name='email']")
    input_pas = driver.find_element(By.CSS_SELECTOR, "input[name='pass']")
    time.sleep(5)

    # вводим логин и пароль
    input_email.clear()
    input_email.send_keys(email)
    time.sleep(random.randrange(3, 5))
    input_pas.clear()
    input_pas.send_keys(pas)
    time.sleep(random.randrange(3, 5))
    # PUSH SUBMIT
    btn = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
    btn.click()
    time.sleep(random.randrange(5, 7))
    #2FA autorize
    if driver.current_url == "https://www.facebook.com/checkpoint/?next":
        print("pause 60 seconds to get 2FA code and click all submit buttons")
        time.sleep(60)

    print("end of timer, go rest of code")
    # go to main account
    driver.get(url="https://www.facebook.com/profile.php?id=100007148140320")
    time.sleep(random.randrange(3, 5))
    # page_scroll
    for j in range(0, 4):
        print("я скролю 4 рази")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.randrange(3, 4))

    time.sleep(5)

    href = driver.find_elements(By.CSS_SELECTOR, "a[role='link']")
    fb_urls = [i.get_attribute("href") for i in href if
               i.get_attribute("href").startswith("https://www.facebook.com/yarosavpugach/posts/")]  # find posts by url
    print(f"get her of posts{fb_urls}")
    FB_data = list()
    for ur in fb_urls[0:8]:
        driver.get(url=ur)
        time.sleep(3)
        repost = driver.find_element(By.XPATH,
                                     "//SPAN[@class='gvxzyvdx aeinzg81 t7p7dqev gh25dzvf exr7barw b6ax4al1 gem102v4 ncib64c9 mrvwc6qr sx8pxkcf f597kf1v cpcgwwas m2nijcs8 hxfwr5lz k1z55t6l oog5qr5w tes86rjd rtxb060y']").text
        comment = driver.find_element(By.XPATH,
                                      "//SPAN[@class='gvxzyvdx aeinzg81 t7p7dqev gh25dzvf exr7barw b6ax4al1 gem102v4 ncib64c9 mrvwc6qr sx8pxkcf f597kf1v cpcgwwas m2nijcs8 hxfwr5lz k1z55t6l oog5qr5w tes86rjd rtxb060y']").text
        likes = driver.find_element(By.XPATH, "//SPAN[@class='cxfqmxzd nnzkd6d7']").text
        comm_n = re.sub('[\D]', '', comment)
        repost_n = re.sub('[\D]', '', repost)
        likes_list = likes.split(",")
        lenth = len(likes_list)
        if lenth >= 1:
            count = int(re.search(r'\d+', likes_list[lenth - 1]).group(0)) + lenth

            likesl = count
        else:
            likesl = likes

        data = {"links": ur, "coment": "-", "likes": likesl, "comments": comm_n, "repost": repost_n}
        FB_data.append(data)
        print(f"data fb: {FB_data}")


    def xlsx( fb_data):
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "ФИО - Пугач Я.Р"
        ws['A3'] = "Город - Киев"
        ws['A5'] = "Телефон - +380953130240"
        ws['A7'] = f"Месяц - {date.today().strftime('%B')}"
        headers = ["Ссылки на публикации инстаграм", "Коментарии", "Количество лайков", "Количество коментариев",
                   "Охват публикаации"]
        ws.append(headers)

        for i in range(0, len(fb_data)):
            dataf = list(fb_data[i].values())
            print(f"Заповнюю фейс даними: {dataf}")
            ws.append(dataf)

        for col in range(1, len(headers) + 1):
            for col in range(1, len(headers) + 1):
                for row in range(8, 15):
                    ws[get_column_letter(col) + f'{row}'].border = Border(left=Side(border_style='thin'),
                                                                          right=Side(border_style='thin'),
                                                                          bottom=Side(border_style='thin'),
                                                                          top=Side(border_style='thin'))

        dim_holder = DimensionHolder(worksheet=ws)
        dim_holder['A8'] = ColumnDimension(ws, width=30)
        for col in range(2, len(headers) + 1):
            dim_holder[get_column_letter(col) + f'{col}'] = ColumnDimension(ws, min=col, max=col, width=22)

        ws.column_dimensions = dim_holder

        wb.save(f"Отчетность_{date.today().strftime('%B')}.xlsx")

    xlsx(FB_data)
except Exception as exp:
    print(exp)
finally:
    driver.close()
    driver.quit()


